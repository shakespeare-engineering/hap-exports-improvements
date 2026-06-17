from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from models.air_system import AirSystem


def safe_number(
    value: float | None
) -> float:
    """
    Convert None to 0.
    """

    return value or 0


def clean_sheet_name(
    name: str
) -> str:
    """
    Remove invalid Excel
    sheet characters.
    """

    invalid_chars = (
        r'[]:*?/\\'
    )

    for char in invalid_chars:
        name = name.replace(
            char,
            "-"
        )

    return name[:31]


def add_section_header(
    sheet,
    row: int,
    title: str
) -> int:
    """
    Add formatted section header.
    """

    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=6
    )

    cell = sheet.cell(
        row=row,
        column=1,
        value=title
    )

    cell.font = Font(
        bold=True
    )

    cell.fill = PatternFill(
        fill_type="solid",
        fgColor="DDDDDD"
    )

    return row + 1


def add_table_column_titles(sheet, row: int, titles: list[str]) -> int:
    """
    Add column titles for the table.
    """

    for i, title in enumerate(titles, start=1):
        sheet.cell(
            row=row,
            column=i,
            value=title
        ).alignment = Alignment(horizontal="center")

    return row + 1


def add_load_row(
    sheet,
    row: int,
    label: str,
    sensible: float | None,
    latent: float | None,
    detail_value: float | None = None,
    detail_units: str | None = None
) -> int:
    """
    Add load row using Excel formulas.
    """

    sheet.cell(
        row=row,
        column=1,
        value=label
    )

    sheet.cell(
        row=row,
        column=2,
        value=safe_number(
            sensible
        )
    )

    sheet.cell(
        row=row,
        column=3,
        value=safe_number(
            latent
        )
    )

    # Total
    sheet.cell(
        row=row,
        column=4,
        value=f"=B{row}+C{row}"
    )

    # Details
    if detail_value is not None:

        detail_text: str = (
            f"{detail_value:,.0f}"
        )

        detail_cell = sheet.cell(
            row=row,
            column=6,
            value=detail_value
        )

        if detail_units:
            detail_cell.number_format = (
                f'#,##0 "{detail_units}"'
            )

        detail_cell.alignment = Alignment(
            horizontal="right"
        )

    return row + 1


def add_subtotal_row(
    sheet,
    row: int,
    title: str,
    start_row: int,
    end_row: int
) -> int:
    """
    Add subtotal row.
    """

    sheet.cell(
        row=row,
        column=1,
        value=title
    )

    sheet.cell(
        row=row,
        column=2,
        value=(
            f"=SUM(B{start_row}:"
            f"B{end_row})"
        )
    )

    sheet.cell(
        row=row,
        column=3,
        value=(
            f"=SUM(C{start_row}:"
            f"C{end_row})"
        )
    )

    sheet.cell(
        row=row,
        column=4,
        value=f"=B{row}+C{row}"
    )

    for col in range(1, 7):  # A through F
        sheet.cell(
            row=row,
            column=col
        ).font = Font(
            bold=True
        )

    return row + 2


def format_sheet(sheet) -> None:
    """
    Apply formatting.
    """

    sheet.column_dimensions[
        "A"
    ].width = 35

    for column in [
        "B",
        "C",
        "D"
    ]:
        sheet.column_dimensions[
            column
        ].width = 14

    sheet.column_dimensions[
        "E"
    ].width = 12

    sheet.column_dimensions[
        "F"
    ].width = 18

    sheet.column_dimensions[
        "K"
    ].width = 30

    for column in [
        "j",
        "k",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R"
    ]:
        sheet.column_dimensions[
            column
        ].width = 14

    sheet.column_dimensions[
        "S"
    ].width = 18

    sheet.column_dimensions[
        "T"
    ].width = 18

    # ======================================
    # Cooling load formatting only
    # ======================================

    for row in range(1, 200):

        # Sensible / Latent / Total
        for column in [
            "B",
            "C",
            "D"
        ]:

            sheet[
                f"{column}{row}"
            ].number_format = (
                '#,##0'
            )

        # Percent column
        sheet[
            f"E{row}"
        ].number_format = (
            '0.0%'
        )


def export_system_checksums(systems: dict[str, AirSystem], output_path: str | Path) -> None:
    """
    Export checksum workbook.
    """

    workbook: Workbook = (Workbook())

    # Remove the default sheet
    workbook.remove(workbook.active)

    # First sheet will be the master sheet
    master_sheet_name = None

    for system in systems.values():

        sheet = (workbook.create_sheet(title=clean_sheet_name(system.name)))

        heat = (system.heat_balance)

        row: int = 1

        # ======================================
        # Title
        # ======================================

        sheet["A1"] = ("System Checksums")

        sheet["A2"] = ("Shakespeare Engineering")

        sheet["A4"] = ("Project")

        sheet["B4"] = (system.project_name)

        sheet["A5"] = ("System Name")

        sheet["B5"] = (system.name)

        sheet["A1"].font = Font(
            bold=True,
            size=16
        )

        sheet["A2"].font = Font(
            italic=True
        )

        # ======================================
        # Global Parameters
        # ======================================

        sheet["E8"] = (
            "Unit Efficacy"
        )

        sheet["F8"] = (
            "CFM/Ton"
        )

        # First sheet becomes master
        if master_sheet_name is None:

            master_sheet_name = (
                sheet.title
            )

            sheet["E9"] = 0.90
            sheet["F9"] = 400

        else:

            sheet["E9"] = (
                f"='{master_sheet_name}'!E9"
            )

            sheet["F9"] = (
                f"='{master_sheet_name}'!F9"
            )

        # Formatting
        sheet["E9"].number_format = (
            '0%'
        )

        sheet["F9"].number_format = (
            '0'
        )

        sheet["E8"].alignment = Alignment(
            horizontal="center"
        )

        sheet["F8"].alignment = Alignment(
            horizontal="center"
        )

        sheet["E8"].fill = PatternFill(
            fill_type="solid",
            fgColor="ffb400"
        )

        sheet["F8"].fill = PatternFill(
            fill_type="solid",
            fgColor="ffb400"
        )

        # ======================================
        # Mechanical Equipment Table
        # ======================================

        equipment_start_col = 11  # K
        equipment_row = 7

        equipment_headers = [
            "Space Name",
            "MBH",
            "% Of Total MBH",
            "CFM",
            "NERFED CFM",
            "OA",
            "Sqft",
            "No. People",
            "Required Outdoor AIR CFM/Person",
            "Required Outdoor AIR CFM/sqft"
        ]

        # Header row
        for col_offset, header in enumerate(
            equipment_headers
        ):

            header_cell = sheet.cell(
                row=equipment_row,
                column=(
                    equipment_start_col
                    + col_offset
                ),
                value=header
            )

            header_cell.font = Font(
                bold=True
            )

            header_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="DDDDDD"
            )

            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # Create a unit size cell based on data
        sheet["J7"] = "Unit Size"
        sheet["J7"].alignment = Alignment(horizontal="center")
        sheet["J7"].fill = PatternFill(
            fill_type="solid",
            fgColor="DDDDDD"
        )
        sheet["J7"].font = Font(bold=True)

        unit_tons_formula = "A57/E9"

        sheet["J8"] = (
            f'=IF({unit_tons_formula}<=0.75,0.75,'
            f'IF({unit_tons_formula}<=1,1,'
            f'IF({unit_tons_formula}<=1.25,1.25,'
            f'IF({unit_tons_formula}<=1.5,1.5,'
            f'IF({unit_tons_formula}<=2,2,'
            f'IF({unit_tons_formula}<=3,3,'
            f'IF({unit_tons_formula}<=4,4,'
            f'IF({unit_tons_formula}<=5,5,'
            f'IF({unit_tons_formula}<=6,6,'
            f'IF({unit_tons_formula}<=7.5,7.5,'
            f'IF({unit_tons_formula}<=8.5,8.5,'
            f'IF({unit_tons_formula}<=10,10,'
            f'IF({unit_tons_formula}<=12.5,12.5,'
            f'IF({unit_tons_formula}<=15,15,'
            f'IF({unit_tons_formula}<=17.5,17.5,'
            f'IF({unit_tons_formula}<=20,20,'
            f'IF({unit_tons_formula}<=25,25,'
            f'IF({unit_tons_formula}<=27.5,27.5,'
            f'{unit_tons_formula}))))))))))))))))))'
        )

        # Create a unit size cell based on data
        sheet["J10"] = "Design CFM"
        sheet["J10"].alignment = Alignment(horizontal="center")
        sheet["J10"].fill = PatternFill(
            fill_type="solid",
            fgColor="DDDDDD"
        )
        sheet["J10"].font = Font(bold=True)
        sheet["J11"] = "=J8*F9"

        equipment_row += 1

        equipment_start_row = equipment_row

        # ======================================
        # Space Rows
        # ======================================

        for zone in system.zones:

            for space in zone.spaces:

                # Mechanical Equipment Name
                sheet.cell(
                    row=equipment_row,
                    column=11,
                    value=space.name
                )

                # MBH
                sheet.cell(
                    row=equipment_row,
                    column=12,
                    value=(
                        safe_number(
                            space.cooling_latent_mbh + space.cooling_sensible_mbh
                        )
                    )
                )

                # % Of Total MBH (Needs calculation after loop)

                # CFM
                sheet.cell(
                    row=equipment_row,
                    column=14,
                    value=f"=J11*M{equipment_row}"
                )

                # Nerfed CFM
                sheet.cell(
                    row=equipment_row,
                    column=15,
                    value=(
                        f"=MROUND("
                        f"N{equipment_row},"
                        f"5)"
                    )
                )

                # OA
                sheet.cell(
                    row=equipment_row,
                    column=16,
                    value=(
                        safe_number(
                            space.uncorrected_oa_cfm
                        )
                    )
                )

                # Sqft
                sheet.cell(
                    row=equipment_row,
                    column=17,
                    value=(
                        safe_number(
                            space.floor_area_sqft
                        )
                    )
                )

                # No. People
                sheet.cell(
                    row=equipment_row,
                    column=18,
                    value=(
                        safe_number(
                            space.maximum_occupants
                        )
                    )
                )

                # Required OA CFM/person
                sheet.cell(
                    row=equipment_row,
                    column=19,
                    value=(
                        safe_number(
                            space.oa_per_person
                        )
                    )
                )

                # Required OA CFM/sqft
                sheet.cell(
                    row=equipment_row,
                    column=20,
                    value=(
                        safe_number(
                            space.oa_per_sqft
                        )
                    )
                )

                equipment_row += 1


        equipment_end_row = equipment_row - 1

        sheet.cell(
            row=equipment_row,
            column=11,   
            value="TOTAL"
        )

        # MBH
        sheet.cell(
            row=equipment_row,
            column=12,   
            value=f"=SUM(L{equipment_start_row}:L{equipment_end_row})"
        )

        # ======================================
        # % of Total MBH formulas
        # ======================================

        for space_row in range(
            equipment_start_row,
            equipment_end_row + 1
        ):

            sheet.cell(
                row=space_row,
                column=13,
                value=(
                    f"=L{space_row}"
                    f"/L{equipment_row}"
                )
            ).number_format = "0.0%"

        # Total row should equal 100%
        sheet.cell(
            row=equipment_row,
            column=13,
            value=1
        )

        sheet.cell(
            row=equipment_row,
            column=13
        ).number_format = "0.0%"

        # %
        sheet.cell(
            row=equipment_row,
            column=13,  
            value=f"=SUM(M{equipment_start_row}:M{equipment_end_row})"
        )

        # CFM
        sheet.cell(
            row=equipment_row,
            column=14,  
            value=f"=SUM(N{equipment_start_row}:N{equipment_end_row})"
        )

        # Nerfed CFM
        sheet.cell(
            row=equipment_row,
            column=15,  
            value=f"=SUM(O{equipment_start_row}:O{equipment_end_row})"
        )

        # OA
        sheet.cell(
            row=equipment_row,
            column=16,  
            value=f"=SUM(P{equipment_start_row}:P{equipment_end_row})"
        )

        # Sqft
        sheet.cell(
            row=equipment_row,
            column=17,  
            value=f"=SUM(Q{equipment_start_row}:Q{equipment_end_row})"
        )

        # People
        sheet.cell(
            row=equipment_row,
            column=18,  
            value=f"=SUM(R{equipment_start_row}:R{equipment_end_row})"
        )

        # Bold the whole row
        for col in range(11, 30):

            sheet.cell(
                row=equipment_row,
                column=col
            ).font = Font(
                bold=True
            )

        # Set the initial row
        row = 7

        # ======================================
        # Cooling Conditions
        # ======================================

        row = (
            add_section_header(
                sheet,
                row,
                "Cooling Conditions"
            )
        )

        sheet.cell(
            row=row,
            column=1,
            value="Peak Time"
        )
        sheet.cell(
            row=row,
            column=2,
            value= heat.cooling_design_time
        )

        row += 1

        sheet.cell(
            row=row,
            column=1,
            value="Outside Air DB"
        )
        sheet.cell(
            row=row,
            column=2,
            value=(heat.cooling_oa_db_f)
        )

        row += 1

        sheet.cell(
            row=row,
            column=1,
            value="Outside Air WB"
        )

        sheet.cell(
            row=row,
            column=2,
            value= heat.cooling_oa_wb_f
        )

        row += 2

        # ======================================
        # Loads Header
        # ======================================

        row = (
            add_section_header(
                sheet,
                row,
                "Cooling Loads"
            )
        )

        titles = [
            "Item",
            "Sensible",
            "Latent",
            "Total Load",
            "% Total Load",
            "Details"
        ]
        row = add_table_column_titles(sheet, row, titles)

        row += 1

        # Keep track of rows that need to be formatted as percentages
        percent_rows: list[int] = []

        # ======================================
        # Envelope Loads
        # ======================================

        row = add_section_header(
            sheet,
            row,
            "Envelope Loads"
        )

        # Keep track of starting row for percentage formatting
        start_row = row

        row = add_load_row(
            sheet,
            row,
            "Wall",
            heat.wall_btu,
            None,
            heat.wall_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Roof",
            heat.roof_btu,
            None,
            heat.roof_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Window",
            heat.window_btu,
            None,
            heat.window_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Skylight",
            heat.skylight_btu,
            None,
            heat.skylight_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Door",
            heat.door_btu,
            None,
            heat.door_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Floor",
            heat.floor_btu,
            None,
            heat.floor_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Interior Wall",
            heat.interior_wall_btu,
            None,
            heat.interior_wall_sqft,
            "sqft"
        )

        row = add_load_row(
            sheet,
            row,
            "Ceiling",
            heat.ceiling_btu,
            None,
            heat.ceiling_sqft,
            "sqft"
        )

        end_row = row - 1

        percent_rows.extend(
            range(
                start_row,
                end_row + 1
            )
        )

        subtotal_row = row

        row = add_subtotal_row(
            sheet,
            row,
            "Envelope Subtotal",
            start_row,
            end_row
        )

        percent_rows.append(
            subtotal_row
        )

        # ======================================
        # Internal Loads
        # ======================================

        row = add_section_header(
            sheet,
            row,
            "Internal Loads"
        )

        start_row = row

        row = add_load_row(
            sheet,
            row,
            "People",
            heat.people_sensible_btu,
            heat.people_latent_btu,
            heat.people_count,
            "people"
        )

        row = add_load_row(
            sheet,
            row,
            "Infiltration",
            heat.infiltration_sensible_btu,
            heat.infiltration_latent_btu,
            heat.infiltration_cfm,
            "CFM"
        )

        row = add_load_row(
            sheet,
            row,
            "Misc Equipment",
            heat.misc_equipment_sensible_btu,
            heat.misc_equipment_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Air Energy Change",
            heat.air_energy_change_sensible_btu,
            heat.air_energy_change_latent_btu
        )

        end_row = row - 1

        percent_rows.extend(
            range(
                start_row,
                end_row + 1
            )
        )

        subtotal_row = row

        row = add_subtotal_row(
            sheet,
            row,
            "Internal Subtotal",
            start_row,
            end_row
        )

        percent_rows.append(
            subtotal_row
        )

        # ======================================
        # Equipment Loads
        # ======================================

        row = add_section_header(
            sheet,
            row,
            "Equipment Loads"
        )

        start_row = row

        row = add_load_row(
            sheet,
            row,
            "Overhead Lighting",
            heat.overhead_lighting_btu,
            None,
            heat.overhead_lighting_watts,
            "W"
        )

        row = add_load_row(
            sheet,
            row,
            "Task Lighting",
            heat.task_lighting_btu,
            None,
            heat.task_lighting_watts,
            "W"
        )

        row = add_load_row(
            sheet,
            row,
            "Equipment",
            heat.equipment_btu,
            None,
            heat.equipment_watts,
            "W"
        )

        end_row = row - 1

        percent_rows.extend(
            range(
                start_row,
                end_row + 1
            )
        )

        subtotal_row = row

        row = add_subtotal_row(
            sheet,
            row,
            "Equipment Subtotal",
            start_row,
            end_row
        )

        percent_rows.append(
            subtotal_row
        )

        # ======================================
        # System Loads
        # ======================================

        row = add_section_header(
            sheet,
            row,
            "System / Airside Loads"
        )

        start_row = row

        row = add_load_row(
            sheet,
            row,
            "Zone Conditioning",
            heat.zone_conditioning_sensible_btu,
            heat.zone_conditioning_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Plenum Load",
            heat.plenum_load_sensible_btu,
            heat.plenum_load_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Return Fan Load",
            heat.return_fan_sensible_btu,
            heat.return_fan_latent_btu,
            heat.return_fan_cfm,
            "CFM"
        )

        row = add_load_row(
            sheet,
            row,
            "Ventilation Load",
            heat.ventilation_sensible_btu,
            heat.ventilation_latent_btu,
            heat.ventilation_cfm,
            "CFM"
        )

        row = add_load_row(
            sheet,
            row,
            "Supply Fan Load",
            heat.supply_fan_sensible_btu,
            heat.supply_fan_latent_btu,
            heat.return_fan_cfm,
            "CFM"
        )

        row = add_load_row(
            sheet,
            row,
            "Zone Fan Coil Fans",
            heat.zone_fan_coils_sensible_btu,
            heat.zone_fan_coils_latent_btu
        )

        end_row = row - 1

        percent_rows.extend(
            range(
                start_row,
                end_row + 1
            )
        )

        subtotal_row = row

        row = add_subtotal_row(
            sheet,
            row,
            "System Load Subtotal",
            start_row,
            end_row
        )

        percent_rows.append(
            subtotal_row
        )

        # ======================================
        # HAP Totals
        # ======================================

        start_row = row


        row = add_section_header(
            sheet,
            row,
            "HAP Totals"
        )

        row = add_load_row(
            sheet,
            row,
            "Total System Loads",
            heat.total_system_sensible_btu,
            heat.total_system_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Central Cooling Coil",
            heat.central_cooling_coil_sensible_btu,
            heat.central_cooling_coil_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Central Heating Coil",
            heat.central_heating_coil_sensible_btu,
            heat.central_heating_coil_latent_btu
        )

        row = add_load_row(
            sheet,
            row,
            "Total Conditioning",
            heat.total_conditioning_sensible_btu,
            heat.total_conditioning_latent_btu
        )

        end_row = row - 1

        percent_rows.extend(
            range(
                start_row,
                end_row + 1
            )
        )

        # ======================================
        # Grand Total
        # ======================================

        sheet.cell(
            row=row,
            column=1,
            value="Grand Total"
        )

        sheet.cell(
            row=row,
            column=4,
            value=f"=D{row - 3}"
        )

        sheet[
            f"A{row}"
        ].font = Font(bold=True)

        sheet[
            f"D{row}"
        ].font = Font(bold=True)

        # ======================================
        # % Total Formulas
        # ======================================

        for percent_row in percent_rows:

            # Skip merged rows
            if (
                sheet.cell(
                    row=percent_row,
                    column=5
                ).__class__.__name__
                == "MergedCell"
            ):
                continue

            sheet.cell(
                row=percent_row,
                column=5,
                value=(
                    f"=D{percent_row}"
                    f"/D{row}"
                )
            )

        # ======================================
        # Engineering Checks
        # ======================================
        row += 2

        row = add_section_header(
            sheet,
            row,
            "Engineering Checks"
        )

        titles = [
            "Tons",
            "cfm/ft^2",
            "cfm/ton",
            "ft^2/ton",
            "No. People",
            "[extra]"
        ]
        row = add_table_column_titles(sheet, row, titles)

        # Tons
        sheet.cell(
            row=row,
            column=1,
            value=f"=D{row-4}/(12*1000)"
        )

        # cfm/ft^2
        sheet.cell(
            row=row,
            column=2,
            value=f"=F{row-13}/F21"
        )

        # cfm/ton
        sheet.cell(
            row=row,
            column=3,
            value=f"=F{row-13}/A{row}"
        )

        # ft^2/ton
        sheet.cell(
            row=row,
            column=4,
            value=f"=F21/(A{row})"
        )

        # No. People
        sheet.cell(
            row=row,
            column=5,
            value="=F27"
        )

        format_sheet(sheet)

        # ======================================
        # Engineering Check Formatting
        # ======================================

        # Tons
        sheet.cell(
            row=row,
            column=1
        ).number_format = (
            '0.00'
        )

        # cfm/ft²
        sheet.cell(
            row=row,
            column=2
        ).number_format = (
            '0.00'
        )

        # cfm/ton
        sheet.cell(
            row=row,
            column=3
        ).number_format = (
            '0'
        )

        # ft²/ton
        sheet.cell(
            row=row,
            column=4
        ).number_format = (
            '0'
        )

        # People
        sheet.cell(
            row=row,
            column=5
        ).number_format = (
            '0'
        )


        
    # ======================================
    # Exportation
    # ======================================

    # Get project name from first system
    project_name = next(
        iter(systems.values())
    ).project_name

    # Clean filename
    invalid_chars = '<>:"/\\|?*'

    for char in invalid_chars:
        project_name = project_name.replace(
            char,
            "-"
        )

    project_name = project_name.strip()

    # Build output filename
    output_path = Path(output_path)

    final_output_path = (
        output_path.parent
        /
        f"{project_name} - System Checksums.xlsx"
    )

    workbook.save(
        final_output_path
    )

    print("\nSaved workbook:")

    print(final_output_path)